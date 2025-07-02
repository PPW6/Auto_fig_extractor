# -*- coding: utf-8 -*-
"""
Created on Tue Aug 22 10:48:04 2023

@author: PPW
"""


import os
import re
import shutil
from PIL import Image
import json
from object_detection.main import run
from scatter import Data_extra
from graph_classification import graph_sure
from paddle_ocr.legend_name import get_color_text
from openpyxl import Workbook, load_workbook

def get_filename(scatter_path):
    FileNum = 0
    txtFileNum = 0
    Filename = []
    # os.listdir(filePath)会读取出当前文件夹下的文件夹和文件
    for file in os.listdir(scatter_path): 
        FileNum += 1 # 统计当前文件夹下的文件夹(不包含子文件夹)和文件的总数
        if file.endswith(".txt"):             
            txtFileNum += 1
            Filename.append(os.path.splitext(file)[0])
#            print(os.path.splitext(file)[0])
#    print(f'-------文件夹下的文件夹和文件总数为:{len(os.listdir(scatter_path))}个---------')
#    print(f'-------文件夹下的txt文件总数为:{txtFileNum}个---------')
    return Filename

def graph_extra(scatter_path, image_path):
    Filename = get_filename(scatter_path)
    for i in range(0, len(Filename)):
        path = os.path.join(scatter_path, Filename[i]+'.png')  
        image = Image.open(path)
        im_width, im_height = image.size
#        print(Filename[i], im_width/im_height)
        if 1.1 < im_width/im_height < 1.8:
            shutil.copy(path, image_path)
            print('not_cut',Filename[i], im_width/im_height)
        else:
            cut_path = os.path.join(cutpath, Filename[i]+'.png')
            shutil.copy(path, cut_path)
            print('need_cut',Filename[i], im_width/im_height)
#    file = os.listdir(cutpath)
#    for n in range(0, len(file)):
#        cut_path = os.path.join(cutpath, file[n])
#        image = Image.open(cut_path)
#        im_width, im_height = image.size
#        if im_width/im_height < 2:               
#           shutil.copy(cut_path, image_path)

def graph_cut(cutpath, image_path):
    file = os.listdir(cutpath)
    for n in range(0, len(file)):
        cut_path = os.path.join(cutpath, file[n])
        search = re.findall('-[a-f]',cut_path)
        if search:               
           shutil.copy(cut_path, image_path)
           
def final_data_extra(img_path_png, ocr_legend_path, ocr_model_path, filename, c_path):
    length = len(json_record_path)    
    all_file = []
    for l in range(0, length):
#    for l in range(14, 18):
        with open(os.path.join(json_record_path[l]+'.json'), 'r', encoding='utf8') as fp:
            json_data = json.load(fp)
            name = filename[l]
        stats = os.stat(os.path.join(json_record_path[l]+'.json'))
        if stats.st_size < 1000000:
            file = {}            
            fc = Data_extra(img_path_png, json_data, name, scatter_path)
            # if 'scatter' in img_path_png:
            #     x_data, y_data, curve_color = fc.save_line()
            if 'line' in img_path_png:
                x_data, y_data, curve_color = fc.save_curve()
            x_data = xy_normal(x_data)
            y_data = xy_normal(y_data)
            graph_name = graph_name_get(name)
            legend_path = os.path.join(image_path,legend_path_get(name))
            dic_color_name = get_color_text(name, legend_path, ocr_legend_path, ocr_model_path, curve_color, c_path)
            doi_ = name.split('-')[0:2]
            doi = "-".join(doi_)
            figure_num = name.split('-')[2]
            file['doi'] = doi
            file['filename'] = name
            file['figure_caption'] = graph_name
            file['x_data'] = x_data
            file['y_data'] = y_data
            file['curve_color'] = curve_color 
            file['legend_name'] = dic_color_name['legend_name']
            file['figure_num'] = figure_num
            if dic_color_name['x_unit'] != None:
                file['x_name'] = dic_color_name['x_name']
                file['x_unit'] = dic_color_name['x_unit']
            else:
                file['x_name'] = dic_color_name['x_name']
                file['x_unit'] = None
            if dic_color_name['y_unit'] != None:
                file['y_name'] = dic_color_name['y_name']
                file['y_unit'] = dic_color_name['y_unit']
            else:
                file['y_name'] = dic_color_name['y_name']
                file['y_unit'] = None
            
            all_file.append(file)
    return all_file


def record2csv(name, excel_file, json_data):
    fig_num = name.split('_')[0]
    curve_num = name
    x_data = json_data['x_data']
    y_data = json_data['y_data']
    doi = json_data['doi']
    legend_name = json_data['legend_name']
    x_name = json_data['x_name']
    x_unit = json_data['x_unit']
    y_name = json_data['y_name']
    y_unit = json_data['y_unit']
    wb = Workbook()
    wb = load_workbook(excel_file)
    sheet = wb.active

    # 获取当前工作表的最后一行
    if sheet.max_row == 1 and sheet.cell(row=1, column=1).value is None:
       start_row = 1  # 空工作表
    else:
       start_row = sheet.max_row + 1

    for i in range(0, len(x_data)):
        current_row = start_row + i
        sheet.cell(row=start_row,   column=1, value=fig_num)
        sheet.cell(row=current_row, column=2, value=curve_num)
        sheet.cell(row=current_row, column=3, value=legend_name)
        sheet.cell(row=current_row, column=4, value=x_data[i])
        sheet.cell(row=current_row, column=5, value=x_name)
        sheet.cell(row=current_row, column=6, value=x_unit)
        sheet.cell(row=current_row, column=7, value=y_data[i])
        sheet.cell(row=current_row, column=8, value=y_name)
        sheet.cell(row=current_row, column=9, value=y_unit)
    wb.save(excel_file)

def graph_name_get(name):
    txt_name = name.split('_')[0]
    graph_name = txt_name.split('-')[0:3]
    graph_name = "-".join(graph_name)
    with open(os.path.join(scatter_path, graph_name+'.txt') ,'r',encoding='utf-8') as file:
        data = file.read()
    return data

def legend_path_get(name):  
    legend_name = name.split('_')[0]
    legend_path = os.path.join(legend_name+'.png')
    return legend_path

def read_record(image_path):
    length = os.listdir(image_path)     
    JSONname = []
    json_record_path = []
    filename = []
    for file in length: 
        if file.endswith(".json"):             
            JSONname.append(os.path.splitext(file)[0])
#            print(JSONname)
    for name in JSONname:
        if 'record' in name:
           record_path = os.path.join(image_path,name)
           json_record_path.append(record_path)
           filename.append(name)
    return json_record_path, filename

def xy_normal(data):
    xy_data = []
    if type(data) == float:
        data_normal = format(data, '.2f')
        xy_data.append(data_normal)
    else:
        for i in range(0,len(data)):
            data_normal = format(data[i], '.2f') #保留两位小数
    #        data_normal = round(data[i]) #四舍五入
            xy_data.append(data_normal)
    return xy_data
    
if __name__ == "__main__":
    PATH = os.getcwd() # 获取当前文件的绝对路径
    scatter_path = ".\\25output_graph"
    image_path = ".\object_detection\\images_key_data\\images7-curve"
    # graph_path = ".\input_graph"
    ocr_model_path = '.\paddle_ocr'
    ocr_legend_path = 'legends'
    # cutpath = os.path.join(scatter_path, 'cut_graph')
    c_path = r".\dictionary\dictionary.ini"
    prop_name = 'UTS' #'conductivity', 'UTS', 'hardness'
#    graph_name = graph_sure(graph_path, prop_name, scatter_path, c_path)#用prop_name选择图片后，需要筛选scatter_path不符合要求图片在跑后面程序
#     print('################## delete graph ##################')
#     input("Please check the %s that is not curve and input 'ok'" % scatter_path)
#     fa = graph_extra(scatter_path, image_path)#需要裁剪cutpath后再跑后面程序
#     print('################## cut graph ##################')
#     input("Please check the %s that needs to be cropped and input 'ok'" % cutpath)
# #    graph_cut(cutpath, image_path)
#     print('################## parse graph ##################')
#    fb = run() #解析单图曲线
    img_scatter_png = ".\object_detection\scatter_line"
    excel_file = 'output.xlsx'
    json_record_path, filename = read_record(image_path)#读取record_json文件
    all_file = final_data_extra(img_scatter_png, ocr_legend_path, ocr_model_path, filename, c_path)
    for l,json_data in enumerate(all_file):
        name = json_data['filename'][:-12]
        record2csv(name, excel_file, json_data)



